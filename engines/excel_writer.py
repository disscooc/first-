# excel_writer.py — Excel报价单
import os,shutil,openpyxl
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border,Side,Alignment
from openpyxl.utils import get_column_letter
import config
from utils import fmt_num,fmt_money,new_path,safe_fn,price_header,comp_remark,stor_remark

def generate(comp, stor, is_total, total_amt, unit, date, company, manager, phone, project):
    tpl = config.EXCEL_TPL_UNIT if is_total else config.EXCEL_TPL_QTY
    if not os.path.exists(tpl): raise FileNotFoundError(f"模板不存在: {tpl}")
    out = new_path(safe_fn(unit), "机时服务报价单", ".xlsx")
    shutil.copy2(tpl, out)
    wb = openpyxl.load_workbook(out)

    # ========== 调试日志 ==========
    DBG = False
    def _log(msg):
        if DBG: print(f"[EXCEL_DBG] {msg}", flush=True)
    _log("=" * 60)
    _log(f"模板: {os.path.basename(tpl)}")
    _log(f"模式: {'单价模板(UNIT)' if tpl == config.EXCEL_TPL_UNIT else '数量模板(QTY)'}")
    _log(f"is_total={is_total}  total_amt={total_amt}")
    _log(f"计算资源 {len(comp)} 项, 存储资源 {len(stor)} 项")
    # ==============================

    try:
        ws = wb.active
        _c(ws,2,2,company); _c(ws,2,4,manager); _c(ws,3,2,project); _c(ws,3,4,phone)
        _c(ws,4,2,unit); _c(ws,4,4,date); _c(ws,7,4,price_header(comp))
        # 保持原有对齐属性，只追加 wrap_text
        old_align = ws.cell(row=7, column=4).alignment
        ws.cell(row=7, column=4).alignment = Alignment(
            horizontal=old_align.horizontal, vertical=old_align.vertical,
            wrap_text=True)
        is_unit_template = tpl == config.EXCEL_TPL_UNIT
        output_col_count = 5 if is_unit_template else 7

        # ========== 调试：模板初始状态 ==========
        _log(f"模板初始: max_row={ws.max_row} max_col={ws.max_column}")
        _log(f"合并单元格: {[str(m) for m in ws.merged_cells.ranges]}")
        for r in range(7, ws.max_row + 1):
            vals = []
            for c in range(1, min(ws.max_column + 1, 9)):
                cv = ws.cell(row=r, column=c).value
                vals.append(str(cv)[:20] if cv is not None else ".")
            _log(f"  模板R{r}: {vals}")
        # ===================================

        # ========== 动态定位模板中的存储分区 ==========
        def _find_storage_header_row(ws, min_row=8):
            """在模板中搜索存储分区标题行（含"存储"关键字）"""
            for r in range(min_row, ws.max_row + 1):
                val = ws.cell(row=r, column=1).value
                if val and isinstance(val, str) and "存储" in val:
                    return r
            return None

        template_storage_row = _find_storage_header_row(ws)
        _log(f"模板存储分区所在行: {template_storage_row}")
        storage_template_row = template_storage_row or 9
        compute_row_format = _capture_row_format(ws, 8, output_col_count)
        storage_section_format = _capture_row_format(ws, storage_template_row, output_col_count)
        storage_header_format = _capture_row_format(ws, storage_template_row + 1, output_col_count)
        storage_data_format = _capture_row_format(ws, storage_template_row + 2, output_col_count)
        total_tax_format = _capture_row_format(ws, storage_template_row + 3, output_col_count)
        total_notax_format = _capture_row_format(ws, storage_template_row + 4, output_col_count)

        nc = max(1,len(comp))
        if nc>1:
            for _ in range(nc-1): _ins(ws,8,9)
        for data_row in range(8, 8 + nc):
            _unmerge_row(ws, data_row)
            _apply_row_format(ws, compute_row_format, data_row)

        row=8
        for i_idx, it in enumerate(comp):
            # 先清空整行旧模板数据，避免残留
            _clear_row_values(ws, row)
            _c(ws,row,1,it.rtype); _c(ws,row,2,it.name); _c(ws,row,3,it.spec)
            # CPU 单价固定3位，其它资源保持原来的2位
            _n(ws,row,4,it.price,_compute_price_format(it))

            # ========== 调试：每行写入详情 ==========
            _log(f"COMP[{i_idx}]: rtype={it.rtype} name={it.name} spec={str(it.spec)[:30]}")
            _log(f"  price={it.price} qty={it.quantity} amount={it.amount}")
            _log(f"  write_row={row}  col1=type col2=name col3=spec col4=price")
            # ======================================

            if is_unit_template:
                _c(ws,row,5,comp_remark(it.rtype))
                _log(f"  col5=remark(is_unit) → '{comp_remark(it.rtype)}'")
            elif is_total:
                _c(ws,row,5,""); _c(ws,row,6,"")
                _log(f"  col5='' col6='' (is_total)")
            else:
                # 数量：整数（无小数）
                _f(ws,row,5,f"=F{row}/D{row}")
                _n(ws,row,6,it.amount,"0.00")
                _c(ws,row,7,comp_remark(it.rtype))
                _log(f"  col5=公式(=F{row}/D{row})  col6={it.amount:0.2f}  col7={comp_remark(it.rtype)}")
            row+=1

        # 计算存储分区起始行：优先用模板中搜索到的位置，否则用公式
        if template_storage_row:
            tr = template_storage_row + max(0, nc - 1)
        else:
            tr = 9 + max(0, nc - 1)
        tpr = tr + 2
        lir = 7 + nc
        _log(f"计算分区: tr={tr} tpr={tpr} lir={lir} (template_storage_row={template_storage_row}, nc={nc})")
        _log(f"存储区域: tr(header)={tr} tpr(data_start)={tpr} lir(last_comp_row)={lir}")

        if not stor:
            ws.delete_rows(tr, tpr-tr+1)
            tor = 8+nc; ner = tor+1
        else:
            hf = any(s.free_quota for s in stor)
            free_qty = fmt_num(_gift_storage_qty(stor), '0.##')
            _apply_row_format(ws, storage_section_format, tr)
            # 清空存储分区标题行旧数据，避免残留
            _clear_row_values(ws, tr)
            _c(ws,tr,1,f"存储资源服务（免费赠送{free_qty}T）" if hf else "存储资源服务")
            _merge_row_range(ws, tr, 1, output_col_count)
            if len(stor)>1:
                for _ in range(len(stor)-1): _ins(ws,tpr,tpr+1)
            sr=tpr
            _unmerge_row(ws, sr - 1)
            _apply_row_format(ws, storage_header_format, sr - 1)
            # 清空存储表头旧数据
            _clear_row_values(ws, sr - 1)
            if is_unit_template:
                _set_storage_header(ws, sr - 1, ["资源类型", "资源名称", "规格", "单价\n(元/TB*月)", "备注"])
            else:
                _set_storage_header(ws, sr - 1, ["资源类型", "资源名称", "规格", "单价\n(元/TB*月)", "使用数量（T）", "金额（元）", "备注"])
            for data_row in range(sr, sr + len(stor)):
                _unmerge_row(ws, data_row)
                _apply_row_format(ws, storage_data_format, data_row)
            for s_idx, it in enumerate(stor):
                # 清空存储数据行旧数据
                _clear_row_values(ws, sr)
                _c(ws,sr,1,it.rtype); _c(ws,sr,2,it.name); _c(ws,sr,3,it.spec)
                # 存储单价：固定2位小数
                _n(ws,sr,4,it.price,"0.00")
                remark_qty = it.quantity if getattr(it, 'free_quota', False) else 0

                # ========== 调试：存储行 ==========
                _log(f"STOR[{s_idx}]: rtype={it.rtype} name={it.name}")
                _log(f"  price={it.price} qty={it.quantity} amount={it.amount} free_quota={getattr(it,'free_quota',False)}")
                _log(f"  write_row={sr}  col1=type col2=name col3=spec col4=price")
                # ================================

                if is_unit_template:
                    _c(ws,sr,5,stor_remark(remark_qty))
                    _log(f"  col5=remark(is_unit) → '{stor_remark(remark_qty)}'")
                else:
                    _n(ws,sr,5,it.quantity,"0.##")
                    _n(ws,sr,6,it.amount,"0.00"); _c(ws,sr,7,stor_remark(remark_qty))
                    _log(f"  col5(qty)={it.quantity} col6(amount)={it.amount:0.2f} col7={stor_remark(remark_qty)}")
                sr+=1
            lir = tpr+len(stor)-1; tor = lir+1; ner = tor+1

        _apply_row_format(ws, total_tax_format, tor)
        _apply_row_format(ws, total_notax_format, ner)
        _restore_total_rows(ws, tor, ner, output_col_count)
        _log(f"合计区域: tor(total_row)={tor} ner(no_tax_row)={ner}")

        if is_unit_template:
            # 单价清单模板：收缩为5列，并填入总价
            _shrink_to_five_columns(ws)
            _n(ws, tor, 3, total_amt, "#,##0.00")
            _n(ws, ner, 3, round(total_amt / config.TAX, 2), "#,##0.00")
            _log(f"  总价写入 col3: {total_amt} (含税)")
        elif is_total:
            _n(ws,tor,3,total_amt,"#,##0.00")
            ws.column_dimensions[get_column_letter(5)].hidden=True
            ws.column_dimensions[get_column_letter(6)].hidden=True
            _log(f"  总价写入 col3: {total_amt} (含税)")
        else:
            _f(ws,tor,3,f"=SUM(F8:F{lir})")
            _log(f"  总价公式 col3: =SUM(F8:F{lir})")
        if not is_unit_template:
            _f(ws,ner,3,f"=C{tor}/1.06")
            _log(f"  不含税公式 col3: =C{tor}/1.06")

        _border(ws, ner, 5 if is_unit_template else 7)
        remark_col = 5 if is_unit_template else 7
        _merge_same_value_blocks(ws, 8, 7 + nc, [1, remark_col])
        wb.save(out)
        _log(f"文件已保存: {out}")
        _log("=" * 60)
        return out
    finally: wb.close()

def _writable_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(row=merged.min_row, column=merged.min_col)
    return cell

def _clear_row_values(ws, row):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None

def _copy_style(src, dst):
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)

def _unmerge_intersections(ws, row, start_col, end_col):
    for merged in list(ws.merged_cells.ranges):
        same_row = merged.min_row <= row <= merged.max_row
        intersects_col = not (merged.max_col < start_col or merged.min_col > end_col)
        if same_row and intersects_col:
            _safe_unmerge(ws, merged)

def _merge_row_range(ws, row, start_col, end_col):
    if end_col <= start_col:
        return
    _unmerge_intersections(ws, row, start_col, end_col)
    src = ws.cell(row=row, column=start_col)
    for col in range(start_col + 1, end_col + 1):
        _copy_style(src, ws.cell(row=row, column=col))
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

def _restore_total_rows(ws, tor, ner, output_col_count):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= tor:
            _safe_unmerge(ws, merged)
    for row in (tor, ner):
        _merge_row_range(ws, row, 1, 2)
        _merge_row_range(ws, row, 3, output_col_count)

def _merge_same_value_blocks(ws, start_row, end_row, columns):
    if end_row <= start_row:
        return
    for col in columns:
        block_start = start_row
        prev_value = _merge_key(ws.cell(row=start_row, column=col).value)
        for row in range(start_row + 1, end_row + 2):
            current_value = _merge_key(ws.cell(row=row, column=col).value) if row <= end_row else None
            if current_value != prev_value:
                if prev_value and row - block_start > 1:
                    _merge_vertical_range(ws, block_start, row - 1, col)
                block_start = row
                prev_value = current_value

def _merge_key(value):
    return str(value or "").strip()

def _merge_vertical_range(ws, start_row, end_row, col):
    if end_row <= start_row:
        return
    _unmerge_intersections(ws, start_row, col, col)
    for row in range(start_row + 1, end_row + 1):
        _unmerge_intersections(ws, row, col, col)
    cell = ws.cell(row=start_row, column=col)
    old_align = cell.alignment
    cell.alignment = Alignment(
        horizontal=old_align.horizontal,
        vertical="center",
        wrap_text=old_align.wrap_text,
        text_rotation=old_align.text_rotation,
        shrink_to_fit=old_align.shrink_to_fit,
        indent=old_align.indent,
    )
    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

def _c(ws,r,c,v): _writable_cell(ws,r,c).value = str(v) if v is not None else ""
def _n(ws,r,c,v,fmt):
    cell=_writable_cell(ws,r,c)
    try: cell.value=float(v) if v else 0
    except: cell.value=0
    cell.number_format=fmt
def _f(ws,r,c,formula): _writable_cell(ws,r,c).value=formula
def _compute_price_format(item):
    return "0.000" if getattr(item, "rtype", "") == "CPU" else "0.00"

def _gift_storage_qty(stor):
    qty = sum((getattr(item, "quantity", 0) or 0) for item in stor if getattr(item, "free_quota", False))
    return qty if qty > 0 else config.FREE_STORAGE

def _ins(ws,sr,dr):
    """插入行并复制模板行的样式和数据"""
    ws.insert_rows(dr)
    _copy_row_dimension(ws, sr, dr)
    for c in range(1,ws.max_column+1):
        src=ws.cell(row=sr,column=c); dst=ws.cell(row=dr,column=c)
        if isinstance(dst, MergedCell):
            continue
        if src.value is not None: dst.value=src.value
        _copy_style(src, dst)

def _copy_row_dimension(ws, sr, dr):
    src = ws.row_dimensions[sr]
    dst = ws.row_dimensions[dr]
    dst.height = src.height
    dst.hidden = src.hidden
    dst.outlineLevel = src.outlineLevel
    dst.collapsed = src.collapsed

def _capture_row_format(ws, row, max_col):
    row_dim = ws.row_dimensions[row]
    cells = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cells.append({
            "font": copy(cell.font),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
            "alignment": copy(cell.alignment),
        })
    return {
        "height": row_dim.height,
        "hidden": row_dim.hidden,
        "outlineLevel": row_dim.outlineLevel,
        "collapsed": row_dim.collapsed,
        "cells": cells,
    }

def _apply_row_format(ws, row_format, row):
    dim = ws.row_dimensions[row]
    dim.height = row_format["height"]
    dim.hidden = row_format["hidden"]
    dim.outlineLevel = row_format["outlineLevel"]
    dim.collapsed = row_format["collapsed"]
    for col, style in enumerate(row_format["cells"], start=1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.font = copy(style["font"])
        cell.border = copy(style["border"])
        cell.fill = copy(style["fill"])
        cell.number_format = style["number_format"]
        cell.protection = copy(style["protection"])
        cell.alignment = copy(style["alignment"])

def _unmerge_row(ws, row):
    ranges = list(ws.merged_cells.ranges)
    for merged in ranges:
        if merged.min_row <= row <= merged.max_row:
            _safe_unmerge(ws, merged)

def _set_storage_header(ws, row, labels):
    for col, label in enumerate(labels, start=1):
        _c(ws, row, col, label)

def _shrink_to_five_columns(ws):
    for merged in list(ws.merged_cells.ranges):
        if merged.max_col > 5:
            _safe_unmerge(ws, merged)
            if merged.min_col <= 5 and (merged.min_col < 5 or merged.min_row < merged.max_row):
                ws.merge_cells(
                    start_row=merged.min_row,
                    start_column=merged.min_col,
                    end_row=merged.max_row,
                    end_column=5,
                )
    if ws.max_column > 5:
        ws.delete_cols(6, ws.max_column - 5)

def _border(ws,lr, max_col=7):
    thin=Side(style="thin"); b=Border(left=thin,right=thin,top=thin,bottom=thin)
    for r in range(7,lr+1):
        for c in range(1,max_col+1):
            cell = ws.cell(row=r,column=c)
            if not isinstance(cell, MergedCell):
                cell.border=b

def _safe_unmerge(ws, merged):
    try:
        ws.unmerge_cells(str(merged))
    except KeyError:
        try:
            ws.merged_cells.ranges.remove(merged)
        except (KeyError, ValueError):
            pass
