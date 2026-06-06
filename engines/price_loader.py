# price_loader.py - load compute resources from List Price workbooks
import glob
import os
import re
import sys
import time

from openpyxl import load_workbook

import config
from models import Resource

GPU_KW = ["NVIDIA", "A800", "A100", "V100", "L20", "RTX", "4090", "3090", "3080", "GPU", "H100", "H20"]
HETERO_KW = ["DCU", "\u5f02\u6784\u52a0\u901f\u5361", "\u5f02\u6784", "K100_AI", "K100"]
CPU_KW = ["CPU", "X86", "AMD", "Intel", "HG"]
VISIBLE_TYPES = {"CPU", "GPU", "\u5f02\u6784\u52a0\u901f\u5361"}
STORAGE_TYPE = "\u5b58\u50a8"

TYPE_KEYS = ["\u8d44\u6e90\u7c7b\u578b", "\u4ea7\u54c1"]
NAME_KEYS = ["\u8d44\u6e90\u540d\u79f0", "\u540d\u79f0"]
REGION_KEYS = ["\u533a\u57df\u540d\u79f0", "\u53ef\u7528\u533a", "\u533a\u57df"]
SPEC_KEYS = ["\u8282\u70b9\u6280\u672f\u89c4\u683c", "\u670d\u52a1\u914d\u7f6e", "\u89c4\u683c"]
SHARED_KEYS = ["\u5171\u4eab\u5355\u4ef7", "\u5355\u4ef7"]
EXCLUSIVE_KEYS = ["\u72ec\u5360\u5355\u4ef7"]


def _norm(raw):
    u = str(raw or "").upper()
    for k in HETERO_KW:
        if k.upper() in u:
            return "\u5f02\u6784\u52a0\u901f\u5361"
    for k in GPU_KW:
        if k.upper() in u:
            return "GPU"
    for k in CPU_KW:
        if k.upper() in u:
            return "CPU"
    return str(raw or "").strip()


def _price(value):
    try:
        text = str(value or "").strip()
        if text in {"", "/", "-", "\u2014"}:
            return 0
        return float(re.sub(r"[^\d.\-]", "", text) or 0)
    except Exception:
        return 0


def _text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _contains_any(text, keys):
    return any(key in text for key in keys)


def find_price():
    files = []
    for directory in config.PRICE_DIRS:
        if not os.path.isdir(directory):
            continue
        files.extend(glob.glob(os.path.join(directory, "**", "*.xlsx"), recursive=True))
    candidates = []
    for path in files:
        name = os.path.basename(path)
        if name.startswith("~$"):
            continue
        if not getattr(sys, "frozen", False):
            parts = set(os.path.normpath(path).split(os.sep))
            if parts.intersection({"build", "dist", "build_share_payload"}):
                continue
        if any(key in name for key in ["List Price", "\u4ef7\u683c", "\u76ee\u5f55", "\u670d\u52a1\u8d44\u6e90"]):
            candidates.append(path)
    return max(candidates, key=os.path.getmtime) if candidates else None


def load(pp=None, timings=None):
    if pp is None:
        pp = find_price()
    if not pp or not os.path.exists(pp):
        return []

    try:
        read_start = time.perf_counter()
        wb = load_workbook(pp, data_only=True)
        if timings is not None:
            timings["read_price_table"] = (time.perf_counter() - read_start) * 1000
    except Exception:
        return []

    parse_start = time.perf_counter()
    try:
        ws = _pick_sheet(wb)
        merged_values = _merged_value_map(ws)
        header_row, columns = _find_header(ws, merged_values)
        if not header_row:
            return []

        resources = []
        last_type = ""
        last_name = ""
        for row in range(header_row + 1, ws.max_row + 1):
            raw_type = _cell_text(ws, row, columns["type"], merged_values)
            name = _cell_text(ws, row, columns["name"], merged_values)
            region = _cell_text(ws, row, columns["region"], merged_values)
            spec = _cell_text(ws, row, columns["spec"], merged_values)
            shared = _cell(ws, row, columns["shared"], merged_values)
            exclusive = _cell(ws, row, columns["exclusive"], merged_values) if columns.get("exclusive") else ""

            if _is_section_end(raw_type, name, region, spec, shared, exclusive, resources):
                break

            if raw_type:
                last_type = raw_type
            effective_type = last_type

            if name:
                last_name = name
            elif spec and last_name:
                name = last_name

            if not spec and not name:
                continue

            norm_type = _norm(f"{effective_type} {name} {spec}")
            if norm_type not in VISIBLE_TYPES:
                continue

            resources.append(
                Resource(
                    raw_type=effective_type,
                    rtype=norm_type,
                    name=name or norm_type,
                    region=region,
                    spec=spec,
                    shared_price=_price(shared),
                    exclusive_price=_price(exclusive),
                )
            )
        return resources
    finally:
        if timings is not None:
            timings["parse_resource_catalog"] = (time.perf_counter() - parse_start) * 1000
        wb.close()


def search(resources, query):
    if not query:
        return list(resources)
    query = query.lower()
    return [
        resource
        for resource in resources
        if query in f"{resource.rtype} {resource.name} {resource.region} {resource.spec}".lower()
    ]


def load_storage(pp=None, timings=None):
    if pp is None:
        pp = find_price()
    if not pp or not os.path.exists(pp):
        return []

    read_start = time.perf_counter()
    try:
        wb = load_workbook(pp, data_only=True)
        if timings is not None:
            timings["read_storage_table"] = (time.perf_counter() - read_start) * 1000
    except Exception:
        return []

    parse_start = time.perf_counter()
    try:
        for ws in wb.worksheets:
            merged_values = _merged_value_map(ws)
            resources = _parse_storage_sheet(ws, merged_values)
            if resources:
                return resources
        return []
    finally:
        if timings is not None:
            timings["parse_storage_catalog"] = (time.perf_counter() - parse_start) * 1000
        wb.close()


def _parse_storage_sheet(ws, merged_values):
    section_row = _find_storage_section_row(ws, merged_values)
    if not section_row:
        return []
    header_row, columns = _find_storage_header(ws, merged_values, section_row)
    if not header_row:
        return []

    resources = []
    last_name = ""
    blank_count = 0
    section_title = "\u5b58\u50a8\u8d44\u6e90\u670d\u52a1"
    for row in range(header_row + 1, ws.max_row + 1):
        first_cell = _cell_text(ws, row, 1, merged_values)
        row_text = _row_text(ws, row, merged_values)
        if _is_storage_section_end(first_cell, row_text):
            break

        name = _cell_text(ws, row, columns["name"], merged_values)
        spec = _cell_text(ws, row, columns["spec"], merged_values)
        region = _cell_text(ws, row, columns.get("region"), merged_values)
        shared = _cell(ws, row, columns.get("shared"), merged_values)
        exclusive = _cell(ws, row, columns.get("exclusive"), merged_values)

        if name:
            last_name = name
        elif spec and last_name:
            name = last_name

        if not name and not spec:
            blank_count += 1
            if blank_count >= 5 and resources:
                break
            continue
        blank_count = 0

        if _looks_like_storage_header(name, spec):
            continue

        resources.append(
            Resource(
                raw_type=section_title or STORAGE_TYPE,
                rtype=STORAGE_TYPE,
                name=name or spec or STORAGE_TYPE,
                region=region,
                spec=spec or name or "",
                shared_price=_price(shared),
                exclusive_price=_price(exclusive),
            )
        )
    return _dedupe_resources(resources)


def _find_storage_section_row(ws, merged_values):
    for row in range(1, ws.max_row + 1):
        text = _row_text(ws, row, merged_values)
        if "\u5b58\u50a8\u8d44\u6e90\u670d\u52a1" in text:
            return row
    return None


def _find_storage_header(ws, merged_values, section_row):
    for row in range(section_row + 1, min(ws.max_row, section_row + 10) + 1):
        cells = [_cell_text(ws, row, column, merged_values) for column in range(1, ws.max_column + 1)]
        name_col = _find_name_column(cells)
        spec_col = _find_column(cells, SPEC_KEYS)
        if name_col and spec_col:
            columns = {
                "name": name_col,
                "spec": spec_col,
                "region": _find_column(cells, REGION_KEYS),
                "shared": _find_column(cells, SHARED_KEYS),
                "exclusive": _find_column(cells, EXCLUSIVE_KEYS),
            }
            return row, columns
    return None, {}


def _row_text(ws, row, merged_values):
    return " ".join(
        _cell_text(ws, row, column, merged_values)
        for column in range(1, ws.max_column + 1)
        if _cell_text(ws, row, column, merged_values)
    )


def _is_storage_section_end(first_cell, row_text):
    if not row_text:
        return False
    end_keys = [
        "\u5f00\u6237\u6bcf\u8d26\u6237",
        "\u4e13\u5c5e\u670d\u52a1\u8282\u70b9",
        "\u8f6f\u4ef6\u53ca\u6570\u636e\u670d\u52a1",
        "\u5176\u4ed6\u670d\u52a1",
        "\u670d\u52a1\u8bf4\u660e",
    ]
    return any(key in first_cell or key in row_text for key in end_keys)


def _looks_like_storage_header(name, spec):
    text = f"{name} {spec}"
    return any(key in text for key in NAME_KEYS) and any(key in text for key in SPEC_KEYS)


def _dedupe_resources(resources):
    result = []
    seen = set()
    for resource in resources:
        key = (resource.name, resource.spec, resource.shared_price)
        if key in seen:
            continue
        seen.add(key)
        result.append(resource)
    return result


def _pick_sheet(wb):
    for sheet_name in wb.sheetnames:
        if "\u8ba1\u7b97" in sheet_name and "\u670d\u52a1" in sheet_name:
            return wb[sheet_name]
    for sheet_name in wb.sheetnames:
        if "\u4ef7\u683c" in sheet_name or "\u670d\u52a1" in sheet_name:
            return wb[sheet_name]
    return wb[wb.sheetnames[0]]


def _merged_value_map(ws):
    values = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, column)] = value
    return values


def _cell(ws, row, column, merged_values):
    if not column:
        return ""
    return merged_values.get((row, column), ws.cell(row, column).value)


def _cell_text(ws, row, column, merged_values):
    return _text(_cell(ws, row, column, merged_values))


def _find_header(ws, merged_values):
    for row in range(1, min(ws.max_row, 30) + 1):
        cells = [_cell_text(ws, row, column, merged_values) for column in range(1, ws.max_column + 1)]
        if not any(_contains_any(cell, TYPE_KEYS) for cell in cells):
            continue
        if not any(_contains_any(cell, SPEC_KEYS) for cell in cells):
            continue

        columns = {
            "type": _find_column(cells, TYPE_KEYS),
            "name": _find_name_column(cells),
            "region": _find_column(cells, REGION_KEYS),
            "spec": _find_column(cells, SPEC_KEYS),
            "shared": _find_column(cells, SHARED_KEYS),
            "exclusive": _find_column(cells, EXCLUSIVE_KEYS),
        }
        if all(columns[key] for key in ["type", "name", "region", "spec", "shared"]):
            return row, columns

    return 3, {"type": 1, "name": 2, "region": 3, "spec": 4, "shared": 5, "exclusive": 6}


def _find_column(cells, keys):
    for index, cell in enumerate(cells, start=1):
        if _contains_any(cell, keys):
            return index
    return None


def _find_name_column(cells):
    for index, cell in enumerate(cells, start=1):
        text = str(cell or "").strip()
        if "\u8d44\u6e90\u540d\u79f0" in text:
            return index
    for index, cell in enumerate(cells, start=1):
        text = str(cell or "").strip()
        if "\u540d\u79f0" in text and not any(key in text for key in REGION_KEYS):
            return index
    return None


def _is_section_end(raw_type, name, region, spec, shared, exclusive, resources):
    if not resources:
        return False
    first_cell = _text(raw_type)
    if not first_cell:
        return False
    end_keys = [
        "\u5f02\u6784\u52a0\u901f\u8282\u70b9",
        "\u5b58\u50a8\u8d44\u6e90\u670d\u52a1",
        "\u5f00\u6237\u6bcf\u8d26\u6237",
        "\u4e13\u5c5e\u670d\u52a1\u8282\u70b9",
        "\u8f6f\u4ef6\u53ca\u6570\u636e\u670d\u52a1",
        "\u5176\u4ed6\u670d\u52a1",
    ]
    if any(key in first_cell for key in end_keys):
        return True
    has_resource_fields = any(_text(value) for value in [name, region, spec, shared, exclusive])
    return not has_resource_fields
